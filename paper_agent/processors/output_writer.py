"""Output writer for analysis results."""
import json
import os
import ast
import re
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from config import Config


# 列表字段统一处理函数
LIST_FIELDS = {"contributions",
               "limitations_reported", "limitations_inferred",
               "research_background", "research_question", "research_objective",
               "model_or_framework",
               "experimental_design", "analysis_process", "relevance_score"}

# 需要证据定位的字段
EVIDENCE_FIELDS = {"research_method", "data_source", "key_findings"}

# 元数据字段
METADATA_FIELDS = {"authors", "year", "journal", "doi", "doi_status"}


def normalize_string_list(value) -> list[str]:
    """将各种输入统一为 list[str]。
    - 已经是列表：直接返回
    - 是字符串形式的列表如 "['a', 'b']"：用 ast.literal_eval 安全解析
    - 其他字符串：包装为单元素列表
    - None / 空：返回空列表
    """
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v) for v in value]
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return []
        try:
            parsed = ast.literal_eval(value)
            if isinstance(parsed, (list, tuple)):
                return [str(v) for v in parsed]
        except (ValueError, SyntaxError):
            pass
        return [value]
    return [str(value)]


def format_list_for_markdown(items: list[str]) -> str:
    """将列表格式化为 Markdown 编号列表"""
    if not items:
        return ""
    return "\n".join(f"{i}. {item}" for i, item in enumerate(items, 1))


def format_list_for_excel(items: list[str]) -> str:
    """将列表格式化为 Excel 单元格内的换行编号列表"""
    if not items:
        return ""
    return "\n".join(f"{i}. {item}" for i, item in enumerate(items, 1))


def _get_evidence_text(value) -> str:
    """从证据字段中提取 text，兼容字符串和 dict 格式"""
    if isinstance(value, dict):
        return value.get("text", "")
    if isinstance(value, str):
        return value
    return ""


def _get_evidence_pages(value) -> list[int]:
    """从证据字段中提取 evidence_pages，兼容字符串和 dict 格式"""
    if isinstance(value, dict):
        pages = value.get("evidence_pages", [])
        if isinstance(pages, list):
            return [int(p) for p in pages if isinstance(p, (int, str)) and str(p).isdigit()]
    return []


def _get_evidence_excerpt(value) -> str:
    """从证据字段中提取 evidence_excerpt"""
    if isinstance(value, dict):
        return value.get("evidence_excerpt", "")
    return ""


def format_evidence_single_for_markdown(value) -> str:
    """格式化单个证据字段（research_method / data_source）为 Markdown"""
    text = _get_evidence_text(value)
    pages = _get_evidence_pages(value)
    excerpt = _get_evidence_excerpt(value)

    lines = [text]
    if pages:
        lines.append(f"   - 证据页码：第 {', '.join(str(p) for p in pages)} 页")
    if excerpt:
        lines.append(f"   - 原文依据：{excerpt}")
    return "\n".join(lines)


def format_evidence_list_for_markdown(items: list[dict]) -> str:
    """格式化列表证据字段（key_findings）为 Markdown 编号列表"""
    if not items:
        return ""
    lines = []
    for i, item in enumerate(items, 1):
        text = _get_evidence_text(item) if isinstance(item, dict) else str(item)
        pages = _get_evidence_pages(item) if isinstance(item, dict) else []
        excerpt = _get_evidence_excerpt(item) if isinstance(item, dict) else ""

        lines.append(f"{i}. {text}")
        if pages:
            lines.append(f"   - 证据页码：第 {', '.join(str(p) for p in pages)} 页")
        if excerpt:
            lines.append(f"   - 原文依据：{excerpt}")
    return "\n".join(lines)


def format_evidence_list_for_excel(items: list[dict]) -> str:
    """格式化列表证据字段为 Excel 单元格内容（只加页码，不保存原文）"""
    if not items:
        return ""
    lines = []
    for i, item in enumerate(items, 1):
        text = _get_evidence_text(item) if isinstance(item, dict) else str(item)
        pages = _get_evidence_pages(item) if isinstance(item, dict) else []

        if pages:
            lines.append(f"{i}. {text}（第{', '.join(str(p) for p in pages)}页）")
        else:
            lines.append(f"{i}. {text}")
    return "\n".join(lines)


class OutputWriter:
    def __init__(self, config: Config):
        self.config = config
        os.makedirs(config.output_dir, exist_ok=True)



        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    def _safe_filename(self, name: str) -> str:
        """清理文件名中不能出现的字符"""
        return re.sub(r'[\\/:*?"<>|]', '-', name)[:100]
    def write(self, result: dict):
        title = result["title"].replace(".pdf", "")
        out_path = Path(self.config.output_dir) / f"{title}.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"Saved: {out_path}")
    def write_all(self, results):
        output_path = os.path.join(self.config.output_dir, f"papers_summary_{self.timestamp}.json")
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)        
        self.write_excel(results)
        for result in results:
            self.write_markdown(result)
    def write_excel(self, results):
        # 先格式化列表字段，避免 pandas 将空列表转为 NaN
        formatted_results = []
        for r in results:
            r_copy = dict(r)
            for col in LIST_FIELDS:
                if col in r_copy:
                    r_copy[col] = format_list_for_excel(normalize_string_list(r_copy[col]))

            # 处理证据字段
            for col in EVIDENCE_FIELDS:
                if col in r_copy:
                    value = r_copy[col]
                    if col == "key_findings":
                        # key_findings 是列表
                        if isinstance(value, list):
                            r_copy[col] = format_evidence_list_for_excel(value)
                        else:
                            r_copy[col] = _get_evidence_text(value)
                    else:
                        # research_method / data_source 是单个
                        if isinstance(value, dict):
                            text = value.get("text", "")
                            pages = value.get("evidence_pages", [])
                            if isinstance(pages, list) and pages:
                                r_copy[col] = f"{text}（第{', '.join(str(p) for p in pages)}页）"
                            else:
                                r_copy[col] = text
                        else:
                            r_copy[col] = _get_evidence_text(value)

            formatted_results.append(r_copy)

        df = pd.DataFrame(formatted_results)
        cols = [
            "original_filename",
            "extracted_title",
            "authors",
            "year",
            "journal",
            "doi",
            "doi_status",
            "research_background",
            "research_question",
            "research_objective",
            "research_method",
            "model_or_framework",
            "data_source",
            "experimental_design",
            "analysis_process",
            "key_findings",
            "contributions",
            "limitations_reported",
            "limitations_inferred",
            "relevance_score",
            "path",
        ]
        cols = [c for c in cols if c in df.columns]
        df = df[cols]

        # 作者列表用中文分号连接
        if "authors" in df.columns:
            df["authors"] = df["authors"].apply(
                lambda v: "；".join(normalize_string_list(v))
            )

        output_path = os.path.join(self.config.output_dir, f"papers_summary_{self.timestamp}.xlsx")
        df.to_excel(output_path, index=False)

        # 调整列宽和格式
        self._format_excel(output_path)

    def _format_excel(self, filepath):
        """自动调整列宽并开启换行，让文字完整显示"""
        wb = load_workbook(filepath)
        ws = wb.active

        # 计算每列最大宽度
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    # 中文字符宽度按 2 算，英文按 1 算
                    value = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in value)
                    max_length = max(max_length, length)
            # 列宽上限 60 避免过宽，最小 10
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 60)

        # 数据行开启自动换行，靠上对齐
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        wb.save(filepath)
    def write_markdown(self, result):
        title = result.get("extracted_title") or result.get("original_filename", "untitled")
        safe_title = self._safe_filename(title)

        # 辅助函数：如果是列表字段则格式化为编号列表，否则直接取字符串
        def fmt(field):
            value = result.get(field, "")
            if field in LIST_FIELDS and value:
                return format_list_for_markdown(normalize_string_list(value))
            return value

        # 格式化作者：列表转字符串
        authors = normalize_string_list(result.get("authors"))
        authors_str = "、".join(authors) if authors else "未识别"

        # 格式化年份和期刊
        year = result.get("year") or "未识别"
        journal = result.get("journal") or "未识别"

        # 格式化 DOI
        doi = result.get("doi") or "—"
        doi_status = result.get("doi_status", "未找到")

        content = f"""# {result.get('extracted_title', title)}

## 一、基本信息

- 作者：{authors_str}
- 年份：{year}
- 期刊/来源：{journal}
- DOI：{doi}
- DOI 状态：{doi_status}

## 研究背景
{fmt('research_background')}

## 研究问题
{fmt('research_question')}

## 研究目标
{fmt('research_objective')}

## 研究方法
{format_evidence_single_for_markdown(result.get('research_method'))}

## 模型或理论框架
{fmt('model_or_framework')}

## 数据来源
{format_evidence_single_for_markdown(result.get('data_source'))}

## 实验或研究设计
{fmt('experimental_design')}

## 分析流程
{fmt('analysis_process')}

## 主要发现
{format_evidence_list_for_markdown(result.get('key_findings'))}

## 研究贡献
{fmt('contributions')}

## 研究局限

### 作者明确提出的局限
{format_list_for_markdown(normalize_string_list(result.get('limitations_reported'))) if normalize_string_list(result.get('limitations_reported')) else '未识别到作者明确陈述的研究局限。'}

### 模型基于论文内容推断的局限
{format_list_for_markdown(normalize_string_list(result.get('limitations_inferred'))) if normalize_string_list(result.get('limitations_inferred')) else '暂无推断内容。'}

> 注意：以上推断内容不是论文作者的直接表述，正式引用前需要核对原文。

## 相关性评分
{fmt('relevance_score')}

## 文件路径
{result.get('path', '')}
"""
        filepath = os.path.join(self.config.output_dir, f"{safe_title}.md")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)

    def write_review_list(self, review_items: list[tuple[dict, list[str]]]):
        """输出人工确认清单。
        review_items: [(result, reasons), ...]
        """
        if not review_items:
            return

        lines = [
            "# 人工确认清单",
            "",
            f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"待确认论文数：{len(review_items)}",
            "",
            "---",
            "",
        ]

        for idx, (result, reasons) in enumerate(review_items, 1):
            lines.append(f"## {idx}. {result.get('original_filename', '未知文件')}")
            lines.append("")
            lines.append("**触发原因：**")
            for reason in reasons:
                lines.append(f"- {reason}")
            lines.append("")

            # 显示原始提取文本或 LLM 响应
            raw_text = result.get("raw_text") or result.get("raw_llm_response") or ""
            if raw_text:
                lines.append("**原始内容（前 500 字符）：**")
                lines.append("```")
                lines.append(raw_text[:500])
                lines.append("```")
            else:
                lines.append("**原始内容：** 无")

            lines.append("")
            lines.append("---")
            lines.append("")

        output_path = os.path.join(self.config.output_dir, "review_list.md")
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        print(f"Saved review list: {output_path}")
